import os
import json
import subprocess
import uuid
import logging
from datetime import datetime
from functools import wraps
import requests
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

db_url = os.environ.get('DATABASE_URL')
if db_url and db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///vexio.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

PORTFOLIO_FILE = os.path.join(os.path.dirname(__file__), 'data', 'portfolio.json')
VISITORS_FILE = os.path.join(os.path.dirname(__file__), 'data', 'visitors.json')

def load_visitors():
    try:
        with open(VISITORS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {'unique_ips': [], 'total': 0}

def save_visitors(data):
    os.makedirs(os.path.dirname(VISITORS_FILE), exist_ok=True)
    with open(VISITORS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.before_request
def track_visitor():
    if request.path.startswith('/static/') or request.path.startswith('/admin') or request.path.startswith('/api/'):
        return
    if session.get('admin_logged_in'):
        return
    ip = request.remote_addr or 'unknown'
    data = load_visitors()
    if ip not in data['unique_ips']:
        data['unique_ips'].append(ip)
        data['total'] = len(data['unique_ips'])
        save_visitors(data)

def load_portfolio():
    try:
        with open(PORTFOLIO_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def save_portfolio(data):
    os.makedirs(os.path.dirname(PORTFOLIO_FILE), exist_ok=True)
    with open(PORTFOLIO_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.after_request
def add_headers(response):
    if response.mimetype in ('text/css', 'application/javascript', 'image/webp', 'image/jpeg', 'image/png', 'image/svg+xml', 'font/woff2', 'font/woff'):
        response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'
    return response

ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = generate_password_hash(os.environ.get('ADMIN_PASSWORD', 'vexio2024'))

TG_TOKEN = os.environ.get('TG_TOKEN', '8591869743:AAFgSoVueO9FXIVjIPgwOnDIPOeeN_5x05s')
TG_ADMIN_ID = os.environ.get('TG_ADMIN_ID', '903104535')

def send_tg_notification(sub):
    text = (
        f"\u2709\ufe0f \u041d\u043e\u0432\u0430\u044f \u0437\u0430\u044f\u0432\u043a\u0430\n"
        f"\U0001F4CB \u041f\u0440\u043e\u0435\u043a\u0442: {sub.project_name}\n"
        f"\U0001F464 \u0418\u043c\u044f: {sub.name}\n"
        f"\U0001F4DE \u0422\u0435\u043b\u0435\u0444\u043e\u043d: {sub.phone}\n"
        f"\U0001F4C1 \u0422\u0438\u043f: {sub.type or '\u2014'}\n"
        f"\U0001F535 \u0421\u0442\u0430\u0442\u0443\u0441: {sub.status}"
    )
    if sub.telegram_username:
        text += f"\n\U0001F916 TG: @{sub.telegram_username}"
    if sub.description:
        text += f"\n\U0001F4DD \u041e\u043f\u0438\u0441\u0430\u043d\u0438\u0435: {sub.description[:200]}"
    if sub.reference:
        text += f"\n\U0001F517 \u0420\u0435\u0444\u0435\u0440\u0435\u043d\u0441: {sub.reference}"
    try:
        resp = requests.post(
            f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage',
            json={'chat_id': TG_ADMIN_ID, 'text': text, 'parse_mode': 'HTML'},
            timeout=10,
        )
        if resp.status_code != 200:
            logger.error(f"TG notify failed: {resp.status_code} {resp.text}")
    except Exception as e:
        logger.error(f"TG notify error: {e}")

class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    project_name = db.Column(db.String(200))
    phone = db.Column(db.String(30))
    type = db.Column(db.String(200))
    description = db.Column(db.Text)
    reference = db.Column(db.String(500), default='')
    catalog = db.Column(db.String(10))
    admin = db.Column(db.String(10))
    telegram = db.Column(db.String(10))
    telegram_username = db.Column(db.String(100))
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'project_id': self.project_id,
            'name': self.name,
            'project_name': self.project_name,
            'phone': self.phone,
            'type': self.type,
            'description': self.description,
            'reference': self.reference,
            'catalog': self.catalog,
            'admin': self.admin,
            'telegram': self.telegram,
            'telegram_username': self.telegram_username,
            'status': self.status or 'pending',
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }

with app.app_context():
    db.create_all()
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        cols = [c['name'] for c in inspector.get_columns('submission')]
        if 'telegram_username' not in cols:
            db.session.execute(db.text('ALTER TABLE submission ADD COLUMN telegram_username VARCHAR(100)'))
            db.session.commit()
            logger.info("Added telegram_username column")
        if 'reference' not in cols:
            db.session.execute(db.text('ALTER TABLE submission ADD COLUMN reference VARCHAR(500) DEFAULT \'\''))
            db.session.commit()
            logger.info("Added reference column")
    except Exception as e:
        logger.warning(f"Migration note: {e}")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

import re

@app.route('/hr')
def hr_redirect():
    return redirect('/hr/'), 301

@app.route('/hr/')
@app.route('/hr/<path:path>')
def serve_hr(path='index.html'):
    return app.send_static_file(f'hr/{path or "index.html"}')

@app.route('/brief')
def brief_redirect():
    return redirect('/brief/'), 301

@app.route('/brief/')
@app.route('/brief/<path:path>')
def serve_brief(path='index.html'):
    return app.send_static_file(f'brief/{path or "index.html"}')

@app.route('/check-files')
def check_files():
    paths = {
        'hr': os.path.join(app.root_path, 'static', 'hr', 'index.html'),
        'brief': os.path.join(app.root_path, 'static', 'brief', 'index.html'),
        'static_hr': os.path.join(app.root_path, 'static', 'hr'),
        'static_brief': os.path.join(app.root_path, 'static', 'brief'),
        'root_path': app.root_path,
        'static_files': os.listdir(os.path.join(app.root_path, 'static'))[:20],
    }
    result = {}
    for k, v in paths.items():
        if k in ('root_path', 'static_files'):
            result[k] = v
        elif os.path.isdir(v):
            result[k] = f'DIR exists, files: {os.listdir(v)[:10]}'
        else:
            result[k] = f'EXISTS ({os.path.getsize(v)} bytes)' if os.path.exists(v) else 'NOT FOUND'
    return jsonify(result)

@app.route('/')
def index():
    return send_from_directory('templates', 'index.html')

@app.route('/favicon.ico')
def favicon():
    svg = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100"><text y=".9em" font-size="90" fill="#8B5CF6">\u2726</text></svg>'
    return app.response_class(svg, mimetype='image/svg+xml')

@app.route('/version')
def version():
    import subprocess, os
    try:
        rev = subprocess.check_output(['git', 'rev-parse', 'HEAD'], cwd=os.path.dirname(__file__)).decode().strip()[:7]
    except: rev = '?'
    host = request.headers.get('X-Forwarded-Host', request.host)
    return f'OK {rev} host={host}'

NOTIFY_FILE = os.path.join(os.path.dirname(__file__), 'data', 'notifications.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'data', 'submissions.xlsx')

def save_notification(source, data):
    os.makedirs(os.path.dirname(NOTIFY_FILE), exist_ok=True)
    try:
        with open(NOTIFY_FILE, 'r', encoding='utf-8') as f:
            queue = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        queue = []
    queue.append({'source': source, 'data': data, 'time': datetime.utcnow().isoformat()})
    with open(NOTIFY_FILE, 'w', encoding='utf-8') as f:
        json.dump(queue, f, ensure_ascii=False)
    _save_to_excel(source, data)

def _save_to_excel(source, data):
    from openpyxl import Workbook, load_workbook
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        wb.remove(wb.active)
    sheet_name = {'site': 'Сайт', 'hr': 'Vexio HR', 'brief': 'Vexio Brief'}.get(source, source)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        if source == 'hr':
            ws.append(['Имя', 'Telegram', 'Телефон', 'Вакансия', 'ID', 'Дата'])
        elif source == 'brief':
            ws.append(['Формат', 'Бизнес', 'Цель', 'Контакт', 'Сайт', 'ID', 'Дата'])
        else:
            ws.append(['Имя', 'Проект', 'Телефон', 'Тип', 'Описание', 'Каталог', 'Админка', 'Telegram бот', 'ID', 'Дата'])
    ws = wb[sheet_name]
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M')
    if source == 'hr':
        ws.append([data.get('name',''), data.get('telegram',''), data.get('phone',''), data.get('position',''), data.get('project_id',''), now])
    elif source == 'brief':
        ws.append([data.get('format',''), data.get('business',''), data.get('goal',''), data.get('contact',''), data.get('website',''), data.get('project_id',''), now])
    else:
        ws.append([data.get('name',''), data.get('project',''), data.get('phone',''), data.get('type',''), data.get('description',''), data.get('catalog','no'), data.get('admin','no'), data.get('telegram','no'), data.get('project_id',''), now])
    wb.save(EXCEL_FILE)

@app.route('/admin/download-excel')
@login_required
def download_excel():
    if not os.path.exists(EXCEL_FILE):
        return 'No submissions yet', 404
    return send_file(EXCEL_FILE, as_attachment=True, download_name='submissions.xlsx')

@app.route('/excel')
def public_excel():
    key = request.args.get('key', '')
    if key != os.environ.get('EXCEL_KEY', 'vexio2024'):
        return 'Unauthorized', 401
    if not os.path.exists(EXCEL_FILE):
        return 'No submissions yet', 404
    return send_file(EXCEL_FILE, as_attachment=True, download_name='submissions.xlsx')

@app.route('/api/notifications')
def get_notifications():
    try:
        with open(NOTIFY_FILE, 'r', encoding='utf-8') as f:
            queue = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        queue = []
    return jsonify(queue)

@app.route('/api/notifications/clear', methods=['POST'])
def clear_notifications():
    os.makedirs(os.path.dirname(NOTIFY_FILE), exist_ok=True)
    with open(NOTIFY_FILE, 'w', encoding='utf-8') as f:
        json.dump([], f)
    return jsonify({'success': True})

@app.route('/api/hr-apply', methods=['POST'])
def hr_apply():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    data['project_id'] = 'HR-' + str(uuid.uuid4())[:6].upper()
    data['source'] = 'hr'
    data['created_at'] = datetime.utcnow().isoformat()
    save_notification('hr', data)
    return jsonify({'success': True})

@app.route('/api/brief-apply', methods=['POST'])
def brief_apply():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    data['project_id'] = 'BF-' + str(uuid.uuid4())[:6].upper()
    data['source'] = 'brief'
    data['created_at'] = datetime.utcnow().isoformat()
    save_notification('brief', data)
    return jsonify({'success': True})


@app.route('/api/ai-chat', methods=['POST'])
def ai_chat():
    data = request.get_json()
    q = (data.get('q') or '').strip()
    if not q: return jsonify({'error':'No question'}), 400
    HF_KEY = os.environ.get('HF_KEY','')
    reply = None
    if HF_KEY:
        try:
            r = requests.post('https://router.huggingface.co/v1/chat/completions',headers={'Authorization':f'Bearer {HF_KEY}'},json={'model':'Qwen/Qwen2.5-7B-Instruct','messages':[{'role':'system','content':'Ты ассистент веб-студии. Отвечай кратко, по-русски.'},{'role':'user','content':q}],'max_tokens':200,'temperature':0.7},timeout=10)
            if r.status_code==200: reply = r.json().get('choices',[{}])[0].get('message',{}).get('content')
        except: pass
    if not reply:
        try:
            r = requests.post('https://text.pollinations.ai/openai',json={'model':'openai','messages':[{'role':'system','content':'Ты ассистент. Отвечай кратко, по-русски.'},{'role':'user','content':q}],'max_tokens':200},timeout=10)
            if r.status_code==200: reply = r.json().get('choices',[{}])[0].get('message',{}).get('content')
        except: pass
    if not reply: reply = rule_answer(q)
    return jsonify({'reply':reply or 'Не знаю ответа'})

def rule_answer(q):
    for k,v in {'цена':'От 15000. Лендинг, от 50000 магазин.','сайт':'Сайты любой сложности. Полный цикл.','бот':'Telegram-боты с CRM, оплатой, AI.','срок':'Лендинг 3-7 дней, сайт 10-20, магазин 14-30.'}.items():
        if k in q.lower(): return v
    return 'Спросите о ценах, сроках, сайтах или ботах!'

@app.route('/news/')
def news_page():
    return send_from_directory('templates', 'news.html')

@app.route('/api/news')
def api_news():
    NEWS_KEY = os.environ.get('NEWS_KEY', '')
    try:
        r = requests.get(
            'https://newsapi.org/v2/top-headlines',
            params={'category': 'technology', 'language': 'en', 'pageSize': 9, 'apiKey': NEWS_KEY},
            timeout=10
        )
        if r.status_code == 200:
            return jsonify(r.json())
    except Exception:
        pass
    return jsonify({'articles': []})

@app.route('/help/')
def help_page():
    return send_from_directory('templates', 'help.html')

@app.route('/api/help-apply', methods=['POST'])
def help_apply():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    save_notification('help', {
        'name': data.get('name', ''),
        'contact': data.get('contact', ''),
        'message': data.get('message', ''),
        'project_id': 'HL-' + str(uuid.uuid4())[:6].upper(),
    })
    return jsonify({'success': True})

@app.route('/api/submit', methods=['POST'])
def submit_project():
    data = request.get_json()
    if not data or not data.get('projectName'):
        return jsonify({'error': 'Project name is required'}), 400
    pid = 'VX-' + str(uuid.uuid4())[:6].upper()
    sub = Submission(
        project_id=pid,
        name=data.get('name', ''),
        project_name=data['projectName'],
        phone=data.get('phone', ''),
        type=data.get('type', ''),
        description=data.get('description', ''),
        reference=data.get('reference', ''),
        catalog=data.get('catalog', 'no'),
        admin=data.get('admin', 'no'),
        telegram=data.get('telegram', 'no'),
        telegram_username=data.get('telegram_username', ''),
    )
    db.session.add(sub)
    db.session.commit()
    save_notification('site', {
        'name': data.get('name', ''),
        'project': data.get('projectName', ''),
        'phone': data.get('phone', ''),
        'type': data.get('type', ''),
        'description': data.get('description', '') or '—',
        'catalog': data.get('catalog', 'no'),
        'admin': data.get('admin', 'no'),
        'telegram': data.get('telegram', 'no'),
    })
    return jsonify({'success': True, 'project_id': pid}), 201

@app.route('/admin', methods=['GET', 'POST'])
def admin_login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            error = 'Invalid credentials'
    return render_template('admin.html', error=error, login=True)

def load_all_submissions():
    subs = [s.to_dict() for s in Submission.query.order_by(Submission.created_at.desc()).all()]
    try:
        with open(NOTIFY_FILE, 'r', encoding='utf-8') as f:
            notify = json.load(f)
        for n in notify:
            d = n['data']
            subs.append({
                'id': d.get('project_id', ''),
                'name': d.get('name', '') or d.get('contact', ''),
                'project_name': d.get('project', '') or d.get('position', '') or d.get('format', ''),
                'phone': d.get('phone', '') or d.get('contact', ''),
                'type': d.get('type', '') or d.get('business', ''),
                'description': d.get('description', '') or d.get('goal', ''),
                'status': 'new',
                'created_at': n.get('time', ''),
                'source': n.get('source', 'site'),
            })
    except (FileNotFoundError, json.JSONDecodeError):
        pass
    subs.sort(key=lambda x: x.get('created_at', ''), reverse=True)
    return subs

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    submissions = load_all_submissions()
    portfolio = load_portfolio()
    visitors = load_visitors()
    return render_template('admin.html', login=False, page='submissions', submissions=submissions, portfolio=portfolio, visitors=visitors)

@app.route('/admin/portfolio', methods=['GET', 'POST'])
@login_required
def admin_portfolio():
    port = load_portfolio()
    if request.method == 'POST':
        category = request.form.get('category', '')
        name = request.form.get('name', '')
        desc = request.form.get('description', '')
        link = request.form.get('link', '').strip()
        image_url = request.form.get('image_url', '').strip()
        if name and category:
            entry = {
                'id': str(uuid.uuid4())[:8],
                'category': category,
                'name': name,
                'description': desc,
                'link': link,
                'image': image_url,
                'created': datetime.utcnow().isoformat(),
            }
            port.append(entry)
            save_portfolio(port)
            if 'image' in request.files:
                file = request.files['image']
                if file.filename:
                    try:
                        r = requests.post(
                            'https://catbox.moe/user/api.php',
                            data={'reqtype': 'fileupload'},
                            files={'fileToUpload': (file.filename, file.stream, file.content_type)},
                            timeout=10,
                        )
                        if r.status_code == 200 and r.text.startswith('https://'):
                            entry['image'] = r.text.strip()
                            save_portfolio(port)
                        else:
                            logger.warning(f"Catbox upload failed: {r.status_code} {r.text[:200]}")
                    except Exception as e:
                        logger.warning(f"Image upload skipped: {e}")
        return redirect(url_for('admin_portfolio'))
    return render_template('admin.html', login=False, page='portfolio', portfolio=port)

@app.route('/admin/portfolio/delete/<item_id>', methods=['POST'])
@login_required
def delete_portfolio_item(item_id):
    port = load_portfolio()
    port = [p for p in port if p['id'] != item_id]
    save_portfolio(port)
    return redirect(url_for('admin_portfolio'))

@app.route('/api/portfolio/<category>')
def api_portfolio(category):
    port = load_portfolio()
    items = [p for p in port if p['category'] == category]
    return jsonify(items)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/api/submissions')
@login_required
def api_submissions():
    submissions = Submission.query.order_by(Submission.created_at.desc()).all()
    return jsonify([s.to_dict() for s in submissions])

@app.route('/admin/api/submissions/<int:sub_id>', methods=['DELETE'])
@login_required
def delete_submission(sub_id):
    sub = Submission.query.get_or_404(sub_id)
    db.session.delete(sub)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/submissions/<int:sub_id>/status', methods=['POST'])
@login_required
def update_submission_status(sub_id):
    sub = Submission.query.get_or_404(sub_id)
    status = request.json.get('status')
    if status not in ('pending', 'accepted', 'rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    sub.status = status
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/referrals')
@login_required
def api_referrals():
    try:
        with open(os.path.join(app.root_path, 'data', 'referrals.json'), 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify([])

@app.route('/api/register_user', methods=['POST'])
def api_register_user():
    data = request.get_json()
    if not data or not data.get('user_id'):
        return jsonify({'error': 'user_id required'}), 400
    users_file = os.path.join(app.root_path, 'data', 'users.json')
    try:
        with open(users_file, 'r', encoding='utf-8') as f:
            users = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        users = []
    uid = data['user_id']
    if not any(u.get('user_id') == uid for u in users):
        users.append({
            'user_id': uid,
            'username': data.get('username', ''),
            'first_name': data.get('first_name', ''),
            'created_at': datetime.utcnow().isoformat(),
        })
        os.makedirs(os.path.dirname(users_file), exist_ok=True)
        with open(users_file, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
    return jsonify({'success': True})

@app.route('/admin/mailing', methods=['GET', 'POST'])
@login_required
def admin_mailing():
    users_file = os.path.join(app.root_path, 'data', 'users.json')
    try:
        with open(users_file, 'r', encoding='utf-8') as f:
            users = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        users = []
    result = None
    if request.method == 'POST':
        text = request.form.get('text', '').strip()
        photo = request.files.get('photo')
        photo_bytes = photo.read() if photo and photo.filename else None
        photo_filename = photo.filename if photo and photo.filename else None
        photo_type = photo.content_type if photo and photo.filename else None
        sent = 0
        failed = 0
        for u in users:
            uid = u.get('user_id')
            if not uid:
                continue
            try:
                if photo_bytes:
                    resp = requests.post(
                        f'https://api.telegram.org/bot{TG_TOKEN}/sendPhoto',
                        files={'photo': (photo_filename, photo_bytes, photo_type)},
                        data={'chat_id': uid, 'caption': text, 'parse_mode': 'HTML'},
                        timeout=15,
                    )
                else:
                    resp = requests.post(
                        f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage',
                        json={'chat_id': uid, 'text': text, 'parse_mode': 'HTML'},
                        timeout=10,
                    )
                if resp.status_code == 200:
                    sent += 1
                else:
                    failed += 1
                    logger.warning(f"Mailing fail to {uid}: {resp.status_code}")
            except Exception as e:
                failed += 1
                logger.error(f"Mailing error to {uid}: {e}")
        result = {'sent': sent, 'failed': failed, 'total': len(users)}
    return render_template('admin.html', login=False, page='mailing', users=users, result=result)

@app.route('/admin/api/users')
@login_required
def api_users():
    try:
        with open(os.path.join(app.root_path, 'data', 'users.json'), 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    except (FileNotFoundError, json.JSONDecodeError):
        return jsonify([])

@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404, message='Page not found'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('error.html', code=500, message='Something went wrong'), 500
